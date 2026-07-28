import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def generate_multi_company_consolidation(dapha_path, rapha_path, output_xlsx_path):
    print("【步驟 1】讀取兩家公司的原始 Excel 訂單資料...")
    df_d = pd.read_excel(dapha_path)
    df_r = pd.read_excel(rapha_path)
    
    # 標記來源公司
    df_d['來源公司'] = '大法'
    df_r['來源公司'] = '拉法'
    
    # 合併資料
    df_all = pd.concat([df_d, df_r], ignore_index=True)
    df_all['地址_clean'] = df_all['地址'].astype(str).str.strip()
    
    print("【步驟 2】跨公司地址整合與併單作業處理中...")
    def consolidate_orders(group):
        info_list = []
        for idx, row in group.iterrows():
            note = str(row['備註']) if pd.notna(row['備註']) else ''
            inv = str(row['發票號碼']) if pd.notna(row['發票號碼']) else ''
            cust_id = str(row['客戶編號']) if pd.notna(row['客戶編號']) else ''
            co = str(row['來源公司'])
            # 貼心加上 [公司名稱] 標籤，一眼看出哪張單是哪家公司的
            info_list.append(f"[{co}] 單號: {row['單據號碼']} | 客編: {cust_id} | 發票: {inv} | 備註: {note}")
        
        原始訊息 = "\n".join(info_list)
        companies = "/".join(group['來源公司'].unique())
        
        return pd.Series({
            '(客戶全稱)': group['(客戶全稱)'].iloc[0],
            '聯絡電話': group['聯絡電話'].iloc[0] if pd.notna(group['聯絡電話'].iloc[0]) else '',
            '地址': group['地址'].iloc[0],
            '來源公司': companies,
            '合併單據數量': len(group),
            '原始訊息': 原始訊息
        })

    # 顯式指定欄位，徹底解決新舊版本 Pandas 相容性與警告問題
    needed_cols = ['(客戶全稱)', '聯絡電話', '地址', '單據號碼', '客戶編號', '發票號碼', '備註', '來源公司']
    consolidated_df = df_all.groupby('地址_clean')[needed_cols].apply(consolidate_orders).reset_index(drop=True)
    
    print("【步驟 3】建立 Excel 多頁籤活頁簿...")
    wb = openpyxl.Workbook()
    
    # 頁籤 1：跨公司統整訂單
    ws_new = wb.active
    ws_new.title = "跨公司統整訂單"
    headers_new = ['(客戶全稱)', '聯絡電話', '地址', '來源公司', '合併單據數量', '原始訊息']
    ws_new.append(headers_new)
    for idx, row in consolidated_df.iterrows():
        ws_new.append([row['(客戶全稱)'], row['聯絡電話'], row['地址'], row['來源公司'], row['合併單據數量'], row['原始訊息']])
        
    # 頁籤 2：大法原始資料
    ws_d = wb.create_sheet(title="大法原始資料")
    ws_d.append(list(df_d.columns[:-1])) # 移除內部臨時欄位
    for r in df_d.itertuples(index=False):
        ws_d.append(list(r)[:-1])
        
    # 頁籤 3：拉法原始資料
    ws_r = wb.create_sheet(title="拉法原始資料")
    ws_r.append(list(df_r.columns[:-1]))
    for r in df_r.itertuples(index=False):
        ws_r.append(list(r)[:-1])
        
    print("【步驟 4】美化報表視覺樣式與寬度自動調整...")
    font_header = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
    font_body = Font(name='Microsoft JhengHei', size=10)
    fill_header = PatternFill(start_color='365F91', end_color='365F91', fill_type='solid')
    fill_zebra = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid')
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    for ws in [ws_new, ws_d, ws_r]:
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 25
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            if ws.title != "跨公司統整訂單":
                ws.row_dimensions[row_idx].height = 20
            for cell in row:
                cell.font = font_body
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = fill_zebra
                
                if ws.title == "跨公司統整訂單" and cell.column == 6:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                elif (ws.title == "跨公司統整訂單" and cell.column in [4, 5]) or (ws.title != "跨公司統整訂單" and cell.column in [1, 3, 6]):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')), 10)
            col_letter = get_column_letter(col[0].column)
            if ws.title == "跨公司統整訂單" and col[0].column == 6:
                ws.column_dimensions[col_letter].width = 75
            else:
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
                
    wb.save(output_xlsx_path)
    print(f"恭喜！跨公司聯合物流併單報表已成功生成：{output_xlsx_path}\n")
    return df_all, consolidated_df

def verify_multi_company_data(df_all, df_conn):
    print("="*50)
    print("【跨公司聯合物流 - 資料驗證程序啟動】")
    print("="*50)
    
    # 驗證 1：兩家公司加總總單數
    orig_total_count = len(df_all)
    conn_sum_count = df_conn['合併單據數量'].sum()
    
    print(f" 兩家公司(大法+拉法)原始總筆數 ：{orig_total_count} 筆")
    print(f" 統整跨公司單據加總數量       ：{conn_sum_count} 筆")
    
    check1 = (orig_total_count == conn_sum_count)
    if check1:
        print(" 檢查 1 結果：[成功] 跨公司總單據數量 100% 吻合，無任何漏單。")
    else:
        print(" 檢查 1 結果：[錯誤] 數量有落差，請勿出貨！")
        
    # 驗證 2：跨公司單號內容反向解析
    all_orig_orders = set(df_all['單據號碼'].astype(str).tolist())
    found_orders = []
    for msg in df_conn['原始訊息'].dropna():
        orders = re.findall(r"單號:\s*(\w+)", msg)
        found_orders.extend(orders)
        
    found_orders_set = set(found_orders)
    missing_orders = all_orig_orders - found_orders_set
    
    check2 = (len(missing_orders) == 0)
    if check2:
        print(" 檢查 2 結果：[成功] 大法與拉法的所有原始單號皆完整存入「原始訊息」欄位中。")
    else:
        print(f" 檢查 2 結果：[錯誤] 遺漏了以下單號：{missing_orders}")
        
    if check1 and check2:
        print("\n 驗證結論：[完美通過] 跨公司資料整合 100% 正確，可以放心進行包裝交寄！")
    else:
        print("\n 驗證結論：[驗證失敗]")
    print("="*50)

if __name__ == "__main__":
    df_all, df_conn = generate_multi_company_consolidation("大法.xls", "拉法.xls", "大法_拉法_聯合物流併單優化.xlsx")
    verify_multi_company_data(df_all, df_conn)
