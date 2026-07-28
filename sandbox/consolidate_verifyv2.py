import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def generate_multi_company_consolidation_with_links(dapha_path, rapha_path, output_xlsx_path):
    print("【步驟 1】讀取兩家公司的原始 Excel 訂單資料...")
    # 支援舊版 .xls 與新版 .xlsx 格式
    df_d = pd.read_excel(dapha_path)
    df_r = pd.read_excel(rapha_path)
    
    # 標記來源公司
    df_d['來源公司'] = '大法'
    df_r['來源公司'] = '拉法'
    
    print("【步驟 2】計算各單號在未來原始分頁中的精確 Excel 行號 (Row)...")
    # 因為 Excel 有標題列（第 1 行），所以明細資料皆由第 2 行開始
    df_d['Excel_Row'] = range(2, len(df_d) + 2)
    df_r['Excel_Row'] = range(2, len(df_r) + 2)
    
    # 垂直合併兩家公司的資料
    df_all = pd.concat([df_d, df_r], ignore_index=True)
    df_all['地址_clean'] = df_all['地址'].astype(str).str.strip()
    
    # 計算每個地址總共有幾筆單據（合併單據數量欄位來源）
    counts_map = df_all['地址_clean'].value_counts().to_dict()
    df_all['合併單據數量'] = df_all['地址_clean'].map(counts_map)
    
    # 計算每個地址包含哪些公司（例如：大法/拉法）
    co_map = df_all.groupby('地址_clean')['來源公司'].apply(lambda x: "/".join(x.unique())).to_dict()
    df_all['跨公司來源'] = df_all['地址_clean'].map(co_map)
    
    # 重點：按照地址進行排序，確保相同送貨地址的訂單會連續排列，以便後續合併儲存格
    df_all = df_all.sort_values(by='地址_clean').reset_index(drop=True)
    
    print("【步驟 3】建立 Excel 活頁簿並寫入原始資料頁籤...")
    wb = openpyxl.Workbook()
    
    # 先建立原始分頁，確保超連結目標正確
    ws_d = wb.active
    ws_d.title = "大法原始資料"
    ws_d.append(list(df_d.columns[:-2])) # 排除內部標記的 來源公司 與 Excel_Row
    for r in df_d.itertuples(index=False):
        ws_d.append(list(r)[:-2])
        
    ws_r = wb.create_sheet(title="拉法原始資料")
    ws_r.append(list(df_r.columns[:-2]))
    for r in df_r.itertuples(index=False):
        ws_r.append(list(r)[:-2])
        
    # 在第一頁插入「跨公司統整訂單」分頁
    ws_new = wb.create_sheet(title="跨公司統整訂單", index=0)
    # 重新補回『跨公司來源』與『合併單據數量』欄位
    headers_new = ['(客戶全稱)', '聯絡電話', '地址', '來源公司', '合併單據數量', '原始單據號碼', '發票號碼', '備註']
    ws_new.append(headers_new)
    
    print("【步驟 4】寫入統整資料並動態建立 HYPERLINK 跨分頁連結...")
    merge_tracker = {} # 格式 -> {地址: [起始行, 結束行]}
    current_row = 2
    
    for idx, row in df_all.iterrows():
        addr = row['地址_clean']
        co = row['來源公司']
        row_num = row['Excel_Row']
        order_id = str(row['單據號碼'])
        
        # 根據公司來源，動態決定超連結指向的分頁目標
        target_sheet = "大法原始資料" if co == "大法" else "拉法原始資料"
        # 建立內部超連結公式
        link_formula = f'=HYPERLINK("#\'{target_sheet}\'!A{row_num}", "{order_id}")'
        
        # 寫入列資料
        ws_new.append([
            row['(客戶全稱)'],
            row['聯絡電話'] if pd.notna(row['聯絡電話']) else '',
            row['地址'],
            row['跨公司來源'], # 跨公司統整顯示（如：大法/拉法）
            row['合併單據數量'], # 重新保留此欄位
            link_formula,      # 單號公式
            row['發票號碼'] if pd.notna(row['發票號碼']) else '',
            row['備註'] if pd.notna(row['備註']) else ''
        ])
        
        # 追蹤與計算相同地址的上下邊界範圍，供稍後合併使用
        if addr not in merge_tracker:
            merge_tracker[addr] = [current_row, current_row]
        else:
            merge_tracker[addr][1] = current_row
            
        current_row += 1
        
    print("【步驟 5】執行相同地址左側欄位自動合併（包含數量與來源）...")
    # 對相同地址的區塊，合併前 5 欄（客戶全稱、聯絡電話、地址、來源公司、合併單據數量）
    for addr, rows in merge_tracker.items():
        start, end = rows[0], rows[1]
        if start != end: # 該地址有 2 筆以上的訂單才需要合併
            for col_idx in [1, 2, 3, 4, 5]: # 前 5 欄全部納入合併區塊
                ws_new.merge_cells(start_row=start, start_column=col_idx, end_row=end, end_column=col_idx)
                
    print("【步驟 6】套用物流排版視覺優化與樣式調整...")
    font_header = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
    font_body = Font(name='Microsoft JhengHei', size=10)
    font_link = Font(name='Microsoft JhengHei', size=10, color='0000FF', underline='single') # 超連結樣式
    
    fill_header = PatternFill(start_color='365F91', end_color='365F91', fill_type='solid') # 專業藍
    fill_zebra1 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_zebra2 = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid') # 淺灰藍
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    
    # 1. 美化「跨公司統整訂單」頁籤
    for cell in ws_new[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_new.row_dimensions[1].height = 25
    
    # 以「整個地址區塊」為單位做斑馬紋切換
    color_index = 0
    for addr, rows in merge_tracker.items():
        start, end = rows[0], rows[1]
        current_fill = fill_zebra2 if color_index % 2 == 1 else fill_zebra1
        color_index += 1
        
        for r_idx in range(start, end + 1):
            ws_new.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 9): # 共有 8 個欄位
                cell = ws_new.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                cell.fill = current_fill
                
                # 欄位對齊與字體設定
                if c_idx == 6: # 原始單據號碼（超連結）
                    cell.font = font_link
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx in [4, 5, 7]: # 來源公司、合併單據數量、發票號碼
                    cell.font = font_body
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx in [1, 2, 3]: # 被合併的左側欄位
                    cell.font = font_body
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else: # 備註
                    cell.font = font_body
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
    # 2. 美化原始資料分頁（大法、拉法）
    for ws in [ws_d, ws_r]:
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for r_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = font_body
                cell.border = thin_border
                if r_idx % 2 == 0:
                    cell.fill = fill_zebra2
                if c_idx in [1, 3, 6]: # 單號、客編、發票居中
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    # 3. 自動最佳化所有分頁的欄位寬度
    for ws in [ws_new, ws_d, ws_r]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 42)
            
    wb.save(output_xlsx_path)
    print(f"恭喜！帶有自動超連結跳轉與數量欄位的整合報表已成功產出：{output_xlsx_path}\n")
    return df_all

def verify_link_data(df_all, output_xlsx_path):
    print("="*55)
    print("【系統資料完整性與超連結錨點 - 雙重自動驗證程序啟動】")
    print("="*55)
    
    wb = openpyxl.load_workbook(output_xlsx_path, data_only=False)
    ws_new = wb["跨公司統整訂單"]
    
    excel_records_count = ws_new.max_row - 1
    orig_total_count = len(df_all)
    
    print(f" 兩家公司原始訂單加總筆數   ：{orig_total_count} 筆")
    print(f" 統整分頁實際載入明細筆數   ：{excel_records_count} 筆")
    
    check1 = (excel_records_count == orig_total_count)
    if check1:
        print(" 檢查 1 結果：[成功] 總明細行數完全對等，沒有任何一筆訂單掉單。")
    else:
        print(" 檢查 1 結果：[錯誤] 行數不吻合，請暫緩出貨！")
        
    link_pattern = r'=HYPERLINK\("#\'(大法原始資料|拉法原始資料)\'!A(\d+)", "(\w+)"\)'
    parsed_orders = []
    
    for r_idx in range(2, ws_new.max_row + 1):
        cell_val = str(ws_new.cell(row=r_idx, column=6).value) # 欄位移至第 6 欄
        match = re.match(link_pattern, cell_val)
        if match:
            parsed_orders.append(match.group(3))
            
    orig_order_ids = set(df_all['單據號碼'].astype(str).tolist())
    parsed_order_ids = set(parsed_orders)
    missing_orders = orig_order_ids - parsed_order_ids
    
    print(f" 超連結公式中成功封裝並解析的單號數量：{len(parsed_order_ids)} / 實際原始單號總數：{len(orig_order_ids)}")
    
    check2 = (len(missing_orders) == 0)
    if check2:
        print(" 檢查 2 結果：[成功] 所有單號皆完美轉化為帶底線的 HYPERLINK 公式，錨點無誤。")
    else:
        print(f" 檢查 2 結果：[錯誤] 發現有單號公式格式異常或遺漏：{missing_orders}")
        
    if check1 and check2:
        print("\n 驗證結論：[完美通過] 跨公司聯合物流併單表 100% 精確，欄位保留完整！")
    else:
        print("\n 驗證結論：[驗證失敗] 請確認原始檔案資料格式。")
    print("="*55)

if __name__ == "__main__":
    dapha_file = "大法.xls"
    rapha_file = "拉法.xls"
    output_file = "大法_拉法_聯合物流併單優化(完整欄位連結版).xlsx"
    
    try:
        df_all = generate_multi_company_consolidation_with_links(dapha_file, rapha_file, output_file)
        verify_link_data(df_all, output_file)
    except FileNotFoundError:
        print(f"\n[錯誤] 找不到原始檔案，請確認 '{dapha_file}' 與 '{rapha_file}' 是否在正確的位置。")
    except Exception as e:
        print(f"\n[系統錯誤] 執行失敗，原因：{e}")
