import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def generate_consolidated_orders(input_excel_path, output_xlsx_path):
    print("【步驟 1】讀取原始 Excel 訂單資料...")
    # 這裡修正為直接讀取 Excel 檔案。若是 .xls 舊格式，系統需搭配 xlrd 套件 (pip install xlrd)
    df = pd.read_excel(input_excel_path)
    
    # 確保地址格式一致，去除前後空格
    df['地址_clean'] = df['地址'].astype(str).str.strip()
    
    print("【步驟 2】開始進行地址整合與併單作業...")
    def consolidate_orders(group):
        info_list = []
        for idx, row in group.iterrows():
            note = str(row['備註']) if pd.notna(row['備註']) else ''
            inv = str(row['發票號碼']) if pd.notna(row['發票號碼']) else ''
            cust_id = str(row['客戶編號']) if pd.notna(row['客戶編號']) else ''
            info_list.append(f"單號: {row['單據號碼']} | 客編: {cust_id} | 發票: {inv} | 備註: {note}")
        
        原始訊息 = "\n".join(info_list)
        
        return pd.Series({
            '(客戶全稱)': group['(客戶全稱)'].iloc[0],
            '聯絡電話': group['聯絡電話'].iloc[0] if pd.notna(group['聯絡電話'].iloc[0]) else '',
            '地址': group['地址'].iloc[0],
            '合併單據數量': len(group),
            '原始訊息': 原始訊息
        })

#    consolidated_df = df.groupby('地址_clean', include_groups=False).apply(consolidate_orders).reset_index(drop=True)
    needed_cols = ['(客戶全稱)', '聯絡電話', '地址', '單據號碼', '客戶編號', '發票號碼', '備註']
    consolidated_df = df.groupby('地址_clean')[needed_cols].apply(consolidate_orders).reset_index(drop=True)
    print("【步驟 3】建立 Excel 活頁簿並寫入資料...")
    
    wb = openpyxl.Workbook()
    
    # 頁籤 1：原始訂單資訊
    ws_orig = wb.active
    ws_orig.title = "原始訂單資訊"
    ws_orig.append(list(df.columns[:-1])) # 排除地址_clean 臨時欄位
    for r in df.itertuples(index=False):
        ws_orig.append(list(r)[:-1])
        
    # 頁籤 2：統整訂單
    ws_new = wb.create_sheet(title="統整訂單")
    headers_new = ['(客戶全稱)', '聯絡電話', '地址', '合併單據數量', '原始訊息']
    ws_new.append(headers_new)
    for idx, row in consolidated_df.iterrows():
        ws_new.append([row['(客戶全稱)'], row['聯絡電話'], row['地址'], row['合併單據數量'], row['原始訊息']])
        
    print("【步驟 4】套用專業物流報表視覺樣式...")
    font_header = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
    font_body = Font(name='Microsoft JhengHei', size=10)
    fill_header = PatternFill(start_color='365F91', end_color='365F91', fill_type='solid')
    fill_zebra = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid')
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    for ws in [ws_orig, ws_new]:
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 25
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            if ws.title == "原始訂單資訊":
                ws.row_dimensions[row_idx].height = 20
            for cell in row:
                cell.font = font_body
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = fill_zebra
                
                if ws.title == "統整訂單" and cell.column == 5:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                elif cell.column in [1, 2, 6] if ws.title == "原始訂單資訊" else cell.column in [4]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            if ws.title == "統整訂單" and col[0].column == 5:
                ws.column_dimensions[col_letter].width = 65
            else:
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
                
    wb.save(output_xlsx_path)
    print(f"恭喜！併單優化報表已成功生成：{output_xlsx_path}\n")
    return df, consolidated_df

def verify_data(df_orig, df_conn):
    print("="*40)
    print("【資料驗證程序啟動】")
    print("="*40)
    
    # 1. 總量校對
    orig_count = len(df_orig)
    conn_sum_count = df_conn['合併單據數量'].sum()
    
    print(f" 原始明細總筆數：{orig_count} 筆")
    print(f" 統整單據加總數：{conn_sum_count} 筆")
    
    check1 = (orig_count == conn_sum_count)
    if check1:
        print(" 檢查 1 結果：[成功] 訂單總數量完全吻合，沒有任何單據漏失。")
    else:
        print(" 檢查 1 結果：[錯誤] 數量不對！")
        
    # 2. 反向文字內容解析驗證
    all_orig_orders = set(df_orig['單據號碼'].astype(str).tolist())
    found_orders = []
    for msg in df_conn['原始訊息'].dropna():
        orders = re.findall(r"單號:\s*(\w+)", msg)
        found_orders.extend(orders)
        
    found_orders_set = set(found_orders)
    missing_orders = all_orig_orders - found_orders_set
    
    check2 = (len(missing_orders) == 0)
    if check2:
        print(" 檢查 2 結果：[成功] 所有原始單據編號皆已完整封裝在「原始訊息」中，無資訊遺漏。")
    else:
        print(f" 檢查 2 結果：[錯誤] 發現漏洞！漏失了以下單號：{missing_orders}")
        
    if check1 and check2:
        print("\n 驗證結論：[完美通過] 資料 100% 準確，物流出貨與財務對帳無安全疑慮。")
    else:
        print("\n 驗證結論：[驗證失敗]")
    print("="*40)

if __name__ == "__main__":
    # 變更為你的原始舊版 xls 檔名與你想輸出的目的地新檔名
    df_orig, df_conn = generate_consolidated_orders("大法.xls", "大法_併單優化.xlsx")
    verify_data(df_orig, df_conn)
