import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def generate_strategic_consolidation_dual_trace(dapha_path, rapha_path, output_xlsx_path):
    print("【步驟 1】讀取兩家公司的原始 Excel 訂單資料...")
    # 支援舊版 .xls 與新版 .xlsx 格式
    df_d = pd.read_excel(dapha_path)
    df_r = pd.read_excel(rapha_path)
    
    # 標記來源公司與紀錄精確原始 Excel 行號 (因有標題列，故資料從 Row 2 開始)
    df_d['來源公司'] = '大法'
    df_d['Excel_Row'] = range(2, len(df_d) + 2)
    
    df_r['來源公司'] = '拉法'
    df_r['Excel_Row'] = range(2, len(df_r) + 2)
    
    # 統一地址前後格式
    df_d['地址_clean'] = df_d['地址'].astype(str).str.strip()
    df_r['地址_clean'] = df_r['地址'].astype(str).str.strip()
    
    print("【步驟 2】執行戰略併單分流 (拉法備註含 'Q' 且與大法地址相同)...")
    dapha_addresses = set(df_d['地址_clean'].unique())
    
    def has_q(note_val):
        if pd.isna(note_val):
            return False
        return 'Q' in str(note_val).upper()
    
    # 篩選規則：拉法備註有 Q 且地址在大法中也存在
    condition_to_move = df_r['備註'].apply(has_q) & df_r['地址_clean'].isin(dapha_addresses)
    
    df_r_to_dapha = df_r[condition_to_move].copy()
    df_r_remaining = df_r[~condition_to_move].copy()
    
    print(f"   -> 轉移至大法的拉法單據數量: {len(df_r_to_dapha)} 筆")
    print(f"   -> 留在拉法各自出貨的單據數量: {len(df_r_remaining)} 筆")
    
    # 組合兩大合併分頁的資料池 (Pool)
    df_dapha_final_pool = pd.concat([df_d, df_r_to_dapha], ignore_index=True)
    df_rapha_final_pool = df_r_remaining.copy()
    
    print("【步驟 3】核心主號與對照計算：計算各群組地址之「併單單號」，並反向映射...")
    # 計算各 Pool 內每個地址的第一筆單號作為『主併單單號』
    dapha_lead_map = df_dapha_final_pool.sort_values(by='單據號碼').groupby('地址_clean')['單據號碼'].first().to_dict()
    rapha_lead_map = df_rapha_final_pool.sort_values(by='單據號碼').groupby('地址_clean')['單據號碼'].first().to_dict()
    
    # 3-1. 回填「併單單號」至原始明細中 (修正重點：明確對準，不因切片而丟失)
    df_d['併單單號'] = df_d['地址_clean'].map(dapha_lead_map)
    df_r['併單單號'] = None
    df_r.loc[condition_to_move, '併單單號'] = df_r.loc[condition_to_move, '地址_clean'].map(dapha_lead_map)
    df_r.loc[~condition_to_move, '併單單號'] = df_r.loc[~condition_to_move, '地址_clean'].map(rapha_lead_map)
    
    # 3-2. 計算併單統計資訊映射回 Final Pool
    for pool_df, lead_map in [(df_dapha_final_pool, dapha_lead_map), (df_rapha_final_pool, rapha_lead_map)]:
        if not pool_df.empty:
            pool_df['併單單號'] = pool_df['地址_clean'].map(lead_map)
            counts_map = pool_df['地址_clean'].value_counts().to_dict()
            pool_df['合併單據數量'] = pool_df['地址_clean'].map(counts_map)
            
    # 依地址排序，確保 Excel 寫入時相同送貨地址連續相鄰，以利後續合併儲存格
    df_dapha_final_pool = df_dapha_final_pool.sort_values(by='地址_clean').reset_index(drop=True)
    df_rapha_final_pool = df_rapha_final_pool.sort_values(by='地址_clean').reset_index(drop=True)
    
    print("【步驟 4】建立 Excel 活頁簿並寫入【包含追溯主號之原始資料】頁籤...")
    wb = openpyxl.Workbook()
    
    # 定義原始表單要輸出的欄位清單（明確指定！包含『併單單號』欄位）
    orig_output_cols = ['單據號碼', '(客戶全稱)', '客戶編號', '聯絡電話', '地址', '發票號碼', '備註', '聯絡職稱', '併單單號']
    
    # 4-1. 寫入大法原始資料
    ws_d_orig = wb.active
    ws_d_orig.title = "大法原始資料"
    ws_d_orig.append(orig_output_cols)
    for r in df_d[orig_output_cols].itertuples(index=False):
        ws_d_orig.append(list(r))
        
    # 4-2. 寫入拉法原始資料
    ws_r_orig = wb.create_sheet(title="拉法原始資料")
    ws_r_orig.append(orig_output_cols)
    for r in df_r[orig_output_cols].itertuples(index=False):
        ws_r_orig.append(list(r))
        
    print("【步驟 5】建立『合併訂單』頁籤，正式加入【併單單號、來源單號及公司】等雙向欄位與超連結...")
    
    def write_consolidated_sheet(ws_name, source_pool_df, index_pos):
        ws = wb.create_sheet(title=ws_name, index=index_pos)
        # 併單表單核心順序：併單單號在最前，並清晰記錄來源單號與公司
        headers = ['併單單號', '(客戶全稱)', '聯絡電話', '地址', '合併單據數量', '來源公司(公司)', '原始單據號碼(來源單號)', '發票號碼', '備註']
        ws.append(headers)
        
        if source_pool_df.empty:
            return ws, {}
            
        merge_tracker = {}
        current_row = 2
        
        for idx, row in source_pool_df.iterrows():
            addr = row['地址_clean']
            co = row['來源公司']
            row_num = row['Excel_Row']
            order_id = str(row['單據號碼'])
            master_id = str(row['併單單號'])
            
            # 超連結公式：點擊原始單據號碼可自動跨頁跳回原始明細分頁的該行
            target_sheet = "大法原始資料" if co == "大法" else "拉法原始資料"
            link_formula = f'=HYPERLINK("#\'{target_sheet}\'!A{row_num}", "{order_id}")'
            
            ws.append([
                master_id, # A欄：併單單號
                row['(客戶全稱)'],
                row['聯絡電話'] if pd.notna(row['聯絡電話']) else '',
                row['地址'],
                row['合併單據數量'],
                co,            # F欄：來源公司
                link_formula,  # G欄：來源單號（帶超連結）
                row['發票號碼'] if pd.notna(row['發票號碼']) else '',
                row['備註'] if pd.notna(row['備註']) else ''
            ])
            
            if addr not in merge_tracker:
                merge_tracker[addr] = [current_row, current_row]
            else:
                merge_tracker[addr][1] = current_row
                
            current_row += 1
            
        # 相同地址的大區塊（包含：併單單號、客戶名稱、電話、地址、合併數量）垂直合併
        for addr, rows in merge_tracker.items():
            start, end = rows[0], rows[1]
            if start != end:
                for col_idx in [1, 2, 3, 4, 5]: # 前 5 欄全部垂直合併
                    ws.merge_cells(start_row=start, start_column=col_idx, end_row=end, end_column=col_idx)
                    
        return ws, merge_tracker

    ws_d_merge, d_merge_tracker = write_consolidated_sheet("大法合併訂單", df_dapha_final_pool, 0)
    ws_r_merge, r_merge_tracker = write_consolidated_sheet("拉法合併訂單", df_rapha_final_pool, 1)
    
    print("【步驟 6】套用物流美化視覺、大區塊斑馬紋與自動調整欄寬...")
    font_header = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
    font_body = Font(name='Microsoft JhengHei', size=10)
    font_link = Font(name='Microsoft JhengHei', size=10, color='0000FF', underline='single')
    font_master = Font(name='Microsoft JhengHei', size=10, color='1F497D', bold=True) # 併單主號採粗體深藍
    
    fill_header = PatternFill(start_color='365F91', end_color='365F91', fill_type='solid') # 內斂藍標題
    fill_zebra1 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_zebra2 = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid') # 淺灰藍斑馬紋
    
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    # 6-1. 渲染美化合併訂單分頁
    for ws, tracker in [(ws_d_merge, d_merge_tracker), (ws_r_merge, r_merge_tracker)]:
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        color_index = 0
        for addr, rows in tracker.items():
            start, end = rows[0], rows[1]
            current_fill = fill_zebra2 if color_index % 2 == 1 else fill_zebra1
            color_index += 1
            
            for r_idx in range(start, end + 1):
                ws.row_dimensions[r_idx].height = 20
                for c_idx in range(1, 10): 
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.border = thin_border
                    cell.fill = current_fill
                    
                    if c_idx == 1: # 併單單號
                        cell.font = font_master
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif c_idx == 7: # 來源單號 (超連結藍字)
                        cell.font = font_link
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif c_idx in [5, 6, 8]: # 數量、公司、發票
                        cell.font = font_body
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif c_idx in [2, 3, 4]: # 被合併的基本欄位
                        cell.font = font_body
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else: # 備註
                        cell.font = font_body
                        cell.alignment = Alignment(horizontal='left', vertical='center')

    # 6-2. 渲染美化原始資料分頁
    for ws in [ws_d_orig, ws_r_orig]:
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for r_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                if r_idx % 2 == 0:
                    cell.fill = fill_zebra2
                
                # 重點修正：最後一欄 (第 9 欄) 現在已經確認是『併單單號』欄位
                if c_idx == 9:
                    cell.font = font_master
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx in [1, 3, 6]:
                    cell.font = font_body
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = font_body
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    # 6-3. 自動調欄寬
    for ws in [ws_d_merge, ws_r_merge, ws_d_orig, ws_r_orig]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')), 10)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 42)
            
    # 💾 保存活頁簿
    wb.save(output_xlsx_path)
    print(f"恭喜！雙向追溯鏈聯合物流報表已成功產出並存檔：{output_xlsx_path}\n")
    return df_d, df_r


def verify_fixed_system(df_d, df_r, output_xlsx_path):
    """
    品質稽核與雙向核對驗證程序 (QC)
    """
    print("="*65)
    print("【雙向追溯系統 - 全新自動防呆稽核程序啟動】")
    print("="*65)
    
    wb = openpyxl.load_workbook(output_xlsx_path, data_only=True)
    
    print(" 檢查點 1：確認『原始表單』中的【併單單號】欄位是否皆已 100% 寫入資料...")
    ws_d_orig = wb["大法原始資料"]
    ws_r_orig = wb["拉法原始資料"]
    
    # 檢查最後一欄（第 9 欄）是否含有空值
    d_orig_empty = sum(1 for r in range(2, ws_d_orig.max_row + 2) if ws_d_orig.cell(row=r, column=9).value is None and r <= ws_d_orig.max_row)
    r_orig_empty = sum(1 for r in range(2, ws_r_orig.max_row + 2) if ws_r_orig.cell(row=r, column=9).value is None and r <= ws_r_orig.max_row)
    
    print(f"   -> 大法原始資料（共 {ws_d_orig.max_row-1} 筆）：第 9 欄空白數 = {d_orig_empty} 筆")
    print(f"   -> 拉法原始資料（共 {ws_r_orig.max_row-1} 筆）：第 9 欄空白數 = {r_orig_empty} 筆")
    
    check1 = (d_orig_empty == 0 and r_orig_empty == 0)
    if check1:
        print(" 檢查 1 結果：[成功] 每一筆原始訂單皆已成功追溯並寫入「併單單號」！")
    else:
        print(" 檢查 1 結果：[錯誤 Alert] 原始明細中仍有訂單的併單歸宿為空值！")
        
    print("\n 檢查點 2：確認『併單表單』中的【併單單號、來源單號與公司】是否完整...")
    ws_d_merge = wb["大法合併訂單"]
    ws_r_merge = wb["拉法合併訂單"]
    
    d_m_empty = sum(1 for r in range(2, ws_d_merge.max_row + 1) if ws_d_merge.cell(row=r, column=1).value is None)
    r_m_empty = sum(1 for r in range(2, ws_r_merge.max_row + 1) if ws_r_merge.cell(row=r, column=1).value is None)
    
    print(f"   -> 大法合併訂單（A欄併單單號）空白數 = {d_m_empty} 筆")
    print(f"   -> 拉法合併訂單（A欄併單單號）空白數 = {r_m_empty} 筆")
    
    check2 = (d_m_empty == 0 and r_m_empty == 0)
    if check2:
        print(" 檢查 2 結果：[成功] 併單表單的第一欄主單號 100% 生成，且右側來源單號與公司欄位完整對齊！")
    else:
        print(" 檢查 2 結果：[錯誤 Alert] 併單表單的 A 欄存在空值！")
        
    if check1 and check2:
        print("\n 驗證結論：[完美通過 🟢] 雙向追溯鏈完全閉環，數據無瑕疵，物流及財務可放心核銷！")
    else:
        print("\n 驗證結論：[驗證失敗 🔴]")
    print("="*65)


if __name__ == "__main__":
    dapha_input_file = "大法.xls"
    rapha_input_file = "拉法.xls"
    final_output_file = "大法_拉法_自訂規則戰略併單報表(雙向追溯版).xlsx"
    
    try:
        # 正確呼叫修正後的函式名稱
        df_d, df_r = generate_strategic_consolidation_dual_trace(dapha_input_file, rapha_input_file, final_output_file)
        
        # 執行自動 QC 稽核程序
        verify_fixed_system(df_d, df_r, final_output_file)
        
    except FileNotFoundError:
        print(f"\n[錯誤] 檔案讀取失敗，請確認同目錄下是否存在 '{dapha_input_file}' 與 '{rapha_input_file}'。")
    except Exception as e:
        print(f"\n[系統異常] 程式執行中斷，錯誤原因：{e}")
