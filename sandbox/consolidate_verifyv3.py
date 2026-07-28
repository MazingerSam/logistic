import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def generate_strategic_consolidation(dapha_path, rapha_path, output_xlsx_path):
    print("【步驟 1】讀取兩家公司的原始 Excel 訂單資料...")
    df_d = pd.read_excel(dapha_path)
    df_r = pd.read_excel(rapha_path)
    
    # 標記來源公司與計算原始精確 Excel 行號 (Row 2 開始)
    df_d['來源公司'] = '大法'
    df_d['Excel_Row'] = range(2, len(df_d) + 2)
    
    df_r['來源公司'] = '拉法'
    df_r['Excel_Row'] = range(2, len(df_r) + 2)
    
    # 清洗與統一地址欄位
    df_d['地址_clean'] = df_d['地址'].astype(str).str.strip()
    df_r['地址_clean'] = df_r['地址'].astype(str).str.strip()
    
    print("【步驟 2】執行新修正併單規則篩選 (拉法備註含 'Q' 且與大法地址相同)...")
    # 建立大法所有存在的不重複地址集合
    # set 是PYTHON集合操作
    dapha_addresses = set(df_d['地址_clean'].unique())
    
    # 判斷拉法訂單中，備註是否有 "Q" (忽略大小寫、去除空白、防空值)
    def has_q(note_val):
        if pd.isna(note_val):
            return False
        return 'Q' in str(note_val).upper()
    
    # 篩選出符合規則 2 的拉法訂單：備註有 Q 且地址在大法中也存在
    condition_to_move = df_r['備註'].apply(has_q) & df_r['地址_clean'].isin(dapha_addresses)
    
    df_r_to_dapha = df_r[condition_to_move].copy()
    df_r_remaining = df_r[~condition_to_move].copy()
    
    print(f"   -> 移轉至大法的拉法訂單: {len(df_r_to_dapha)} 筆")
    print(f"   -> 留在拉法各自併單的訂單: {len(df_r_remaining)} 筆")
    
    # 分流組合：大法合併訂單分頁資料來源
    df_dapha_final_pool = pd.concat([df_d, df_r_to_dapha], ignore_index=True)
    # 分流組合：拉法合併訂單分頁資料來源
    df_rapha_final_pool = df_r_remaining.copy()
    
    # 預先計算各自 pool 內每個地址的「合併單據數量」與「跨公司來源」
    for pool_df in [df_dapha_final_pool, df_rapha_final_pool]:
        if not pool_df.empty:
            counts_map = pool_df['地址_clean'].value_counts().to_dict()
            co_map = pool_df.groupby('地址_clean')['來源公司'].apply(lambda x: "/".join(x.unique())).to_dict()
            pool_df['合併單據數量'] = pool_df['地址_clean'].map(counts_map)
            pool_df['跨公司來源'] = pool_df['地址_clean'].map(co_map)
            
    # 按照地址排序，確保寫入 Excel 時相同地址相鄰以進行儲存格合併
    df_dapha_final_pool = df_dapha_final_pool.sort_values(by='地址_clean').reset_index(drop=True)
    df_rapha_final_pool = df_rapha_final_pool.sort_values(by='地址_clean').reset_index(drop=True)
    
    print("【步驟 3】建立 Excel 活頁簿與寫入原始資料頁籤...")
    wb = openpyxl.Workbook()
    
    # 先建立後方原始資料分頁
    ws_d_orig = wb.active
    ws_d_orig.title = "大法原始資料"
    ws_d_orig.append(list(df_d.columns[:-3])) # 排除內部加工欄位
    for r in df_d.itertuples(index=False):
        ws_d_orig.append(list(r)[:-3])
        
    ws_r_orig = wb.create_sheet(title="拉法原始資料")
    ws_r_orig.append(list(df_r.columns[:-3]))
    for r in df_r.itertuples(index=False):
        ws_r_orig.append(list(r)[:-3])
        
    print("【步驟 4】建立並寫入『大法合併訂單』與『拉法合併訂單』並內嵌超連結...")
    
    # 定義寫入與合併儲存格的共用邏輯函數
    def write_consolidated_sheet(ws_name, source_pool_df, index_pos):
        ws = wb.create_sheet(title=ws_name, index=index_pos)
        headers = ['(客戶全稱)', '聯絡電話', '地址', '來源公司', '合併單據數量', '原始單據號碼', '發票號碼', '備註']
        ws.append(headers)
        
        if source_pool_df.empty:
            return ws, {}
            
        merge_tracker = {}
        current_row = 2
        
        for idx, row in source_pool_df.iterrows():
            addr = row['地址_clean']
            co = row['來源公司']
            row_num = row['Excel_Row']
            order_id = str(row['單據號碼'])
            
            # 動態錨點超連結指向原始分頁
            target_sheet = "大法原始資料" if co == "大法" else "拉法原始資料"
            link_formula = f'=HYPERLINK("#\'{target_sheet}\'!A{row_num}", "{order_id}")'
            
            ws.append([
                row['(客戶全稱)'],
                row['聯絡電話'] if pd.notna(row['聯絡電話']) else '',
                row['地址'],
                row['跨公司來源'],
                row['合併單據數量'],
                link_formula,
                row['發票號碼'] if pd.notna(row['發票號碼']) else '',
                row['備註'] if pd.notna(row['備註']) else ''
            ])
            
            if addr not in merge_tracker:
                merge_tracker[addr] = [current_row, current_row]
            else:
                merge_tracker[addr][1] = current_row
                
            current_row += 1
            
        # 自動執行前 5 欄儲存格合併
        for addr, rows in merge_tracker.items():
            start, end = rows[0], rows[1]
            if start != end:
                for col_idx in [1, 2, 3, 4, 5]:
                    ws.merge_cells(start_row=start, start_column=col_idx, end_row=end, end_column=col_idx)
                    
        return ws, merge_tracker

    # 建立兩大合併分頁
    ws_d_merge, d_merge_tracker = write_consolidated_sheet("大法合併訂單", df_dapha_final_pool, 0)
    ws_r_merge, r_merge_tracker = write_consolidated_sheet("拉法合併訂單", df_rapha_final_pool, 1)
    
    print("【步驟 5】套用物流美化視覺、斑馬紋與自動調整欄寬...")
    # 樣式宣告
    font_header = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
    font_body = Font(name='Microsoft JhengHei', size=10)
    font_link = Font(name='Microsoft JhengHei', size=10, color='0000FF', underline='single')
    
    fill_header = PatternFill(start_color='365F91', end_color='365F91', fill_type='solid')
    fill_zebra1 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_zebra2 = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid')
    
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    # 渲染合併分頁 (以地址區塊為單位套用斑馬紋)
    for ws, tracker in [(ws_d_merge, d_merge_tracker), (ws_r_merge, r_merge_tracker)]:
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        color_index = 0
        for addr, rows in tracker.items():
            start, end = rows[0], rows[1]
            current_fill = fill_zebra2 if color_index % 2 == 1 else fill_zebra1
            color_index += 1
            
            for r_idx in range(start, end + 1):
                ws.row_dimensions[r_idx].height = 20
                for c_idx in range(1, 9):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.border = thin_border
                    cell.fill = current_fill
                    
                    if c_idx == 6: # 單號超連結
                        cell.font = font_link
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif c_idx in [4, 5, 7]: # 來源、數量、發票
                        cell.font = font_body
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif c_idx in [1, 2, 3]: # 被合併的左側資訊
                        cell.font = font_body
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.font = font_body
                        cell.alignment = Alignment(horizontal='left', vertical='center')

    # 渲染原始分頁
    for ws in [ws_d_orig, ws_r_orig]:
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for r_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = font_body
                cell.border = thin_border
                if r_idx % 2 == 0:
                    cell.fill = fill_zebra2
                if c_idx in [1, 3, 6]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    # 自動調欄寬
    for ws in [ws_d_merge, ws_r_merge, ws_d_orig, ws_r_orig]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 42)
            
    wb.save(output_xlsx_path)
    print(f"恭喜！符合全新規則與超連結的聯合物流報表已生成：{output_xlsx_path}\n")
    return df_d, df_r, df_dapha_final_pool, df_rapha_final_pool

def verify_strategic_data(df_d_orig, df_r_orig, output_xlsx_path):
    print("="*60)
    print("【新物流規則 - 資料完整性與流向驗證程序啟動】")
    print("="*60)
    
    wb = openpyxl.load_workbook(output_xlsx_path, data_only=False)
    
    # 讀取生成後的兩個核心合併分頁
    ws_d_merge = wb["大法合併訂單"]
    ws_r_merge = wb["拉法合併訂單"]
    
    # 計算各自寫入的單據總明細行數
    d_merge_count = ws_d_merge.max_row - 1
    r_merge_count = ws_r_merge.max_row - 1
    excel_total_sum = d_merge_count + r_merge_count
    
    orig_total_sum = len(df_d_orig) + len(df_r_orig)
    
    print(f" 原始訂單總量 (大法: {len(df_d_orig)} 筆 / 拉法: {len(df_r_orig)} 筆) \t= 總計 {orig_total_sum} 筆")
    print(f" 產出 Excel 總量 (大法分頁: {d_merge_count} 筆 / 拉法分頁: {r_merge_count} 筆) \t= 總計 {excel_total_sum} 筆")
    
    # 驗證 1：總數是否不增不減
    check1 = (orig_total_sum == excel_total_sum)
    if check1:
        print(" 檢查 1 結果：[成功] 分流併單後的總單據張數完全吻合，沒有任何訂單掉單！")
    else:
        print(" 檢查 1 結果：[錯誤] 總數不對，資料有外洩或漏失，請勿出貨！")
        
    # 驗證 2：反向單號追蹤與跨分頁錨點正確性
    link_pattern = r'=HYPERLINK\("#\'(大法原始資料|拉法原始資料)\'!A\d+", "(\w+)"\)'
    all_orig_order_ids = set(df_d_orig['單據號碼'].astype(str).tolist() + df_r_orig['單據號碼'].astype(str).tolist())
    
    parsed_order_ids = set()
    for ws in [ws_d_merge, ws_r_merge]:
        for r_idx in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(row=r_idx, column=6).value) # 單號在第 6 欄 (F欄)
            match = re.match(link_pattern, cell_val)
            if match:
                parsed_order_ids.add(match.group(2))
                
    missing_orders = all_orig_order_ids - parsed_order_ids
    print(f" 成功轉化並解析為 HYPERLINK 的單號數：{len(parsed_order_ids)} / 應處理單號總數：{len(all_orig_order_ids)}")
    
    check2 = (len(missing_orders) == 0)
    if check2:
        print(" 檢查 2 結果：[成功] 所有流向不同分頁的單號皆已正確轉為對應的原始檔案跳轉超連結。")
    else:
        print(f" 檢查 2 結果：[錯誤] 發現漏失或公式格式錯誤的單號：{missing_orders}")
        
    if check1 and check2:
        print("\n 驗證結論：[完美通過] 戰略併單優化完成。資料 100% 精確，物流與財務皆可放心對帳！")
    else:
        print("\n 驗證結論：[驗證失敗]")
    print("="*60)

if __name__ == "__main__":
    # 本地測試路徑設定
    dapha_in = "大法.xls"
    rapha_in = "拉法.xls"
    output_out = "大法_拉法_自訂規則戰略併單報表.xlsx"
    
    try:
        df_d, df_r, _, _ = generate_strategic_consolidation(dapha_in, rapha_in, output_out)
        verify_strategic_data(df_d, df_r, output_out)
    except FileNotFoundError:
        print(f"\n[錯誤] 找不到原始 Excel 檔案，請確認 '{dapha_in}' 與 '{rapha_in}' 是否放置於同目錄下。")
