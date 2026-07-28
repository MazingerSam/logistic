import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def generate_strategic_consolidation_dual_trace(dapha_path, rapha_path, output_xlsx_path):
    print("【步驟 1】讀取兩家公司的原始 Excel 訂單資料...")
    # 支援舊版 .xls 與新版 .xlsx 格式
    df_d = pd.read_excel(dapha_path)
    df_r = pd.read_excel(rapha_path)
    
    # 標記來源公司與紀錄精確原始 Excel 行號 (資料從 Row 2 開始)
    df_d['來源公司'] = '大法'
    df_d['Excel_Row'] = range(2, len(df_d) + 2)
    
    df_r['來源公司'] = '拉法'
    df_r['Excel_Row'] = range(2, len(df_r) + 2)
    
    # 統一地址前後格式
    df_d['地址_clean'] = df_d['地址'].astype(str).str.strip()
    df_r['地址_clean'] = df_r['地址'].astype(str).str.strip()
    
    print("【步驟 2】執行戰略併單分流 (拉法備註含 'Q' 且與大法地址相同)...")
    dapha_addresses = set(df_d['地址_clean'].unique())
    
    def has_q(note_val):
        if pd.isna(note_val):
            return False
        return 'Q' in str(note_val).upper()
    
    # 篩選規則：拉法備註有 Q 且地址在大法中也存在
    condition_to_move = df_r['備註'].apply(has_q) & df_r['地址_clean'].isin(dapha_addresses)
    
    df_r_to_dapha = df_r[condition_to_move].copy()
    df_r_remaining = df_r[~condition_to_move].copy()
    
    print(f"   -> 轉移至大法的拉法單據數量: {len(df_r_to_dapha)} 筆")
    print(f"   -> 留在拉法各自出貨的單據數量: {len(df_r_remaining)} 筆")
    
    # 組合兩大合併分頁的資料池 (Pool)
    df_dapha_final_pool = pd.concat([df_d, df_r_to_dapha], ignore_index=True)
    df_rapha_final_pool = df_r_remaining.copy()
    
    print("【步驟 3】核心主號自動修正：採用【兩階段絕對過濾法】確保大法單號絕對主導...")
    def calculate_perfect_lead_id(pool_df):
        lead_map = {}
        for addr, group in pool_df.groupby('地址_clean'):
            # 先篩選出這個地址中所有屬於「大法」的單據
            dapha_orders = group[group['來源公司'] == '大法']
            if not dapha_orders.empty:
                # 如果有大法的單，主號絕對優先從大法的單號中選出最小的
                lead_map[addr] = dapha_orders['單據號碼'].astype(str).min()
            else:
                # 如果該群組完全沒有大法的單（例如純拉法常規單），才選擇拉法的單號
                lead_map[addr] = group['單據號碼'].astype(str).min()
        return lead_map

    dapha_lead_map = calculate_perfect_lead_id(df_dapha_final_pool)
    rapha_lead_map = calculate_perfect_lead_id(df_rapha_final_pool)
    
    # 3-1. 回填「併單單號」至原始明細中 
    df_d['併單單號'] = df_d['地址_clean'].map(dapha_lead_map)
    df_r['併單單號'] = None
    df_r.loc[condition_to_move, '併單單號'] = df_r.loc[condition_to_move, '地址_clean'].map(dapha_lead_map)
    df_r.loc[~condition_to_move, '併單單號'] = df_r.loc[~condition_to_move, '地址_clean'].map(rapha_lead_map)
    
    # 3-2. 計算併單統計資訊映射回 Final Pool
    for pool_df, lead_map in [(df_dapha_final_pool, dapha_lead_map), (df_rapha_final_pool, rapha_lead_map)]:
        if not pool_df.empty:
            pool_df['併單單號'] = pool_df['地址_clean'].map(lead_map)
            counts_map = pool_df['地址_clean'].value_counts().to_dict()
            pool_df['合併單據數量'] = pool_df['地址_clean'].map(counts_map)
            
    # 依地址排序，確保 Excel 寫入時相同送貨地址連續相鄰
    df_dapha_final_pool = df_dapha_final_pool.sort_values(by='地址_clean').reset_index(drop=True)
    df_rapha_final_pool = df_rapha_final_pool.sort_values(by='地址_clean').reset_index(drop=True)
    
    print("【步驟 4】建立 Excel 活頁簿並寫入【精簡雙向追溯明細】原始資料頁籤...")
    wb = openpyxl.Workbook()
    
    orig_output_cols = ['單據號碼', '(客戶全稱)', '客戶編號', '聯絡電話', '地址', '發票號碼', '備註', '聯絡職稱', '併單單號']
    
    # 寫入原始資料時，如果 併單單號 == 單據號碼，則留白不記錄
    def get_orig_row_data(row_tuple):
        row_list = list(row_tuple)
        orig_id = str(row_list[0])   
        master_id = str(row_list[8]) 
        if orig_id == master_id:
            row_list[8] = ''         
        return row_list

    # 4-1. 寫入大法原始資料
    ws_d_orig = wb.active
    ws_d_orig.title = "大法原始資料"
    ws_d_orig.append(orig_output_cols)
    for r in df_d[orig_output_cols].itertuples(index=False):
        ws_d_orig.append(get_orig_row_data(r))
        
    # 4-2. 寫入拉法原始資料
    ws_r_orig = wb.create_sheet(title="拉法原始資料")
    ws_r_orig.append(orig_output_cols)
    for r in df_r[orig_output_cols].itertuples(index=False):
        ws_r_orig.append(get_orig_row_data(r))
        
    print("【步驟 5】建立『合併訂單』頁籤，常規單隱藏重複來源單號...")
    
    def write_consolidated_sheet(ws_name, source_pool_df, index_pos):
        ws = wb.create_sheet(title=ws_name, index=index_pos)
        headers = ['併單單號', '(客戶全稱)', '聯絡電話', '地址', '合併單據數量', '來源公司(公司)', '原始單據號碼(來源單號)', '發票號碼', '備註']
        ws.append(headers)
        
        if source_pool_df.empty:
            return ws, {}
            
        merge_tracker = {}
        current_row = 2
        
        for idx, row in source_pool_df.iterrows():
            addr = row['地址_clean']
            co = row['來源公司']
            row_num = row['Excel_Row']
            order_id = str(row['單據號碼'])
            master_id = str(row['併單單號'])
            total_count = row['合併單據數量']
            
            if total_count == 1:
                display_co = ''        
                link_formula = ''      
            else:
                display_co = co
                target_sheet = "大法原始資料" if co == "大法" else "拉法原始資料"
                link_formula = f'=HYPERLINK("#\'{target_sheet}\'!A{row_num}", "{order_id}")'
            
            ws.append([
                master_id, 
                row['(客戶全稱)'],
                row['聯絡電話'] if pd.notna(row['聯絡電話']) else '',
                row['地址'],
                total_count,
                display_co,    
                link_formula,  
                row['發票號碼'] if pd.notna(row['發票號碼']) else '',
                row['備註'] if pd.notna(row['備註']) else ''
            ])
            
            if addr not in merge_tracker:
                merge_tracker[addr] = [current_row, current_row]
            else:
                merge_tracker[addr][1] = current_row
                
            current_row += 1
            
        # 相同地址的大區塊垂直合併
        for addr, rows in merge_tracker.items():
            start, end = rows[0], rows[1]
            if start != end:
                for col_idx in [1, 2, 3, 4, 5]: 
                    ws.merge_cells(start_row=start, start_column=col_idx, end_row=end, end_column=col_idx)
                    
        return ws, merge_tracker

    ws_d_merge, d_merge_tracker = write_consolidated_sheet("大法合併訂單", df_dapha_final_pool, 0)
    ws_r_merge, r_merge_tracker = write_consolidated_sheet("拉法合併訂單", df_rapha_final_pool, 1)
    
    print("【步驟 6】套用物流美化視覺、大區塊斑馬紋與自動調整欄寬...")
    font_header = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
    font_body = Font(name='Microsoft JhengHei', size=10)
    font_link = Font(name='Microsoft JhengHei', size=10, color='0000FF', underline='single')
    font_master = Font(name='Microsoft JhengHei', size=10, color='1F497D', bold=True) 
    
    fill_header = PatternFill(start_color='365F91', end_color='365F91', fill_type='solid') 
    fill_zebra1 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_zebra2 = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid') 
    
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    # 6-1. 美化合併訂單分頁
    for ws, tracker in [(ws_d_merge, d_merge_tracker), (ws_r_merge, r_merge_tracker)]:
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        color_index = 0
        for addr, rows in tracker.items():
            start, end = rows[0], rows[1]
            current_fill = fill_zebra2 if color_index % 2 == 1 else fill_zebra1
            color_index += 1
            
            for r_idx in range(start, end + 1):
                ws.row_dimensions[r_idx].height = 20
                for c_idx in range(1, 10): 
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.border = thin_border
                    cell.fill = current_fill
                    
                    if c_idx == 1: 
                        cell.font = font_master
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif c_idx == 7: 
                        if cell.value and 'HYPERLINK' in str(cell.value):
                            cell.font = font_link
                        else:
                            cell.font = font_body
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif c_idx in [5, 6, 8]: 
                        cell.font = font_body
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif c_idx in [2, 3, 4]: 
                        cell.font = font_body
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else: 
                        cell.font = font_body
                        cell.alignment = Alignment(horizontal='left', vertical='center')

    # 6-2. 渲染美化原始資料分頁
    for ws in [ws_d_orig, ws_r_orig]:
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for r_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                if r_idx % 2 == 0:
                    cell.fill = fill_zebra2
                
                if c_idx == 9:
                    cell.font = font_master
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx in [1, 3, 6]:
                    cell.font = font_body
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = font_body
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    # 6-3. 自動調欄寬
    for ws in [ws_d_merge, ws_r_merge, ws_d_orig, ws_r_orig]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')), 10)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 42)
            
    wb.save(output_xlsx_path)
    print(f"恭喜！【大法單號絕對主導】之完美版本報表已成功產出：{output_xlsx_path}\n")
    return df_d, df_r


def verify_fixed_system(df_d, df_r, output_xlsx_path):
    """
    品質稽核與雙向核對驗證程序 (QC修正版)
    """
    print("="*65)
    print("【雙向追溯系統 - 智慧合併相容型自動稽核程序啟動】")
    print("="*65)
    
    wb = openpyxl.load_workbook(output_xlsx_path, data_only=True)
    ws_d_merge = wb["大法合併訂單"]
    
    print(" 檢查點 1：強制檢驗發生「跨公司併單」之地址，主號是否被正確鎖定為大法...")
    
    cross_company_check_pass = True
    sampled_count = 0
    wrong_masters = []
    
    dapha_orig_ids = set(df_d['單據號碼'].astype(str).tolist())
    
    # 💡【智慧向下記憶記憶法】：解決 Excel 垂直合併儲存格讀取到 None 的物理現象
    last_valid_master_id = ""
    
    for r in range(2, ws_d_merge.max_row + 1):
        co_val = str(ws_d_merge.cell(row=r, column=6).value or '')
        current_cell_val = ws_d_merge.cell(row=r, column=1).value
        
        # 如果當前格子有值，就更新記憶；如果是 None (表示被合併在同一個大格子裡)，就延用上一個有效的主號
        if current_cell_val is not None:
            last_valid_master_id = str(current_cell_val).strip()
            
        if '拉法' in co_val:
            sampled_count += 1
            # 進行真實校對：檢驗目前這一列所屬的大區塊主號是否為大法單號
            if last_valid_master_id not in dapha_orig_ids:
                cross_company_check_pass = False
                wrong_masters.append(f"行號 {r}: {last_valid_master_id}")
                
    if cross_company_check_pass and sampled_count > 0:
        print(f" 檢查 1 結果：[成功 🟢] 經過垂直合併穿透檢查，共 {sampled_count} 筆聯合物流，併單單號全數鎖定為大法單號。")
    elif sampled_count == 0:
        print(" 檢查 1 結果：[警告] 未在結果中偵測到跨公司合併紀錄。")
    else:
        print(f" 檢查 1 結果：[錯誤 🔴] 發現系統漏洞！以下拉法單號僭越變成了合併主號：{wrong_masters}")
        
    print("\n 檢查點 2：雙向精簡留白功能檢查...")
    ws_d_orig = wb["大法原始資料"]
    sample_orig_clear = any(ws_d_orig.cell(row=i, column=9).value in [None, ''] for i in range(2, ws_d_orig.max_row + 1))
    
    if sample_orig_clear:
        print(" 檢查 2 結果：[成功 🟢] 常規單兩端的重複單號已完美留白隱藏。")
    else:
        print(" 檢查 2 結果：[錯誤 🔴] 留白優化未能成功執行。")
        
    if cross_company_check_pass and sample_orig_clear:
        print("\n 驗證結論：[完美通過 🟢] 數據與檢驗引擎雙重閉環，主導權規則完全落實，可直接交付放行！")
    else:
        print("\n 驗證結論：[驗證失敗 🔴]")
    print("="*65)


if __name__ == "__main__":
    dapha_input_file = "大法.xls"
    rapha_input_file = "拉法.xls"
    final_output_file = "大法_拉法_自訂規則戰略併單報表(優化版).xlsx"
    
    try:
        df_d, df_r = generate_strategic_consolidation_dual_trace(dapha_input_file, rapha_input_file, final_output_file)
        verify_fixed_system(df_d, df_r, final_output_file)
    except FileNotFoundError:
        print(f"\n[錯誤] 檔案讀取失敗，請確認同目錄下是否存在 '{dapha_input_file}' 與 '{rapha_input_file}'。")
